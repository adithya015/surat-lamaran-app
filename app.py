import base64
from io import BytesIO
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import pandas as pd
import streamlit as st
from xhtml2pdf import pisa

# Safe Import PDF Merger
try:
  from pypdf import PdfMerger
except ImportError:
  try:
    from PyPDF2 import PdfMerger
  except ImportError:
    PdfMerger = None

st.set_page_config(
    page_title="Generator Surat & CV", page_icon="📝", layout="centered"
)

# Session State untuk menampung preview
if "pdf_ready" not in st.session_state:
  st.session_state.pdf_ready = False
if "final_pdf_data" not in st.session_state:
  st.session_state.final_pdf_data = None
if "nama_pdf" not in st.session_state:
  st.session_state.nama_pdf = ""
if "email_tujuan" not in st.session_state:
  st.session_state.email_tujuan = ""
if "subjek_email" not in st.session_state:
  st.session_state.subjek_email = ""
if "body_email" not in st.session_state:
  st.session_state.body_email = ""

tab1, tab2 = st.tabs(
    ["📝 Buat Surat & Preview Email", "📊 Rekap Data Lamaran"]
)

# Data Profil Pelamar (Sesuai CV)
nama_pelamar = "Adithya Marhaendra Kusuma"
alamat_pelamar = "Perumahan Mutiara Citra Asri Blok D4 No 6 Candi Sidoarjo"
telepon = "082131009200"
email_pelamar = "adit.marhaendra@gmail.com"
pendidikan = "S2 Teknologi Informasi - ISTTS"
file_excel = "data_lamaran.xlsx"
GMAIL_APP_PASSWORD_DEFAULT = "ssav gwlb tjwz exph"


def dapatkan_html_ttd():
  opsi = [
      "ttd_hd_white.png",
      "ttd_hd_transparent.png",
      "ttd.png",
      "ttd.jpg",
      "ttd.PNG",
  ]
  for f in opsi:
    if os.path.exists(f):
      with open(f, "rb") as img_file:
        b64 = base64.b64encode(img_file.read()).decode()
      return f'<img src="data:image/png;base64,{b64}" style="max-height: 55px; margin: 2px 0;">'
  return '<div style="height:55px;">[Tanda Tangan]</div>'


# ================= TAB 1: FORM SURAT & PREVIEW EMAIL =================
with tab1:
  st.subheader("1. Detail Lamaran Kerja")

  uploaded_cv = st.file_uploader(
      "📄 Upload File CV PDF Baru (Opsional)", type=["pdf"]
  )

  with st.form("form_surat"):
    perusahaan = st.text_input(
        "Nama Perusahaan", value="PT Mayora Indah Tbk (Plant Purwosari)"
    )
    lokasi = st.text_input("Lokasi / Kota", value="Purwosari, Pasuruan")
    posisi = st.text_input("Posisi Dilamar", value="Warehouse Unit Head")
    tanggal = st.text_input("Tanggal Surat", value="29 Agustus 2026")

    st.divider()
    st.subheader("2. Tujuan Pengiriman Email")

    email_tujuan = st.text_input(
        "Email Tujuan (HRD)", value="RECRUITMENT.PWS@MAYORA.CO.ID"
    )
    domisili_subjek = st.text_input(
        "Domisili Anda (Untuk Subjek Email)", value="SIDOARJO"
    )

    gmail_app_password = st.text_input(
        "🔑 Gmail App Password",
        value=GMAIL_APP_PASSWORD_DEFAULT,
        type="password",
    )

    submit_generate = st.form_submit_button(
        "🔍 Generate PDF & Preview Email"
    )

  if submit_generate:
    if not perusahaan or not posisi:
      st.error("Mohon isi Nama Perusahaan dan Posisi!")
    else:
      # Format Subjek Otomatis
      posisi_code = (
          "UH WAREHOUSE" if "WAREHOUSE" in posisi.upper() else posisi.upper()
      )
      subjek_draft = f"{posisi_code}_{domisili_subjek.upper()}"

      # Body Email Dinamis & Spesifik Keterampilan CV (Sari Roti, SCM, Best SCM 2025, IT)
      body_draft = f"""Yth. Tim Recruitment / HRD {perusahaan}
di {lokasi}

Dengan hormat,

Sehubungan dengan informasi lowongan pekerjaan yang dibuka oleh {perusahaan}, melalui email ini saya bermaksud untuk mengajukan diri guna mengisi posisi {posisi}.

Saya merupakan lulusan S2 Teknologi Informasi dari ISTTS (IPK 3.88) dan S1 Teknik Informatika STT PLN dengan pengalaman kerja lebih dari 8 tahun di bidang Supply Chain Management (SCM), Data Control, serta IT & Application Support pada industri FMCG. Saat ini saya menjabat sebagai Subsection SCM Head Control Tower (SPV) di PT Indosari Niaga Nusantara (Sari Roti Group), dengan fokus keahlian:
- Monitoring operasional gudang DC, distributor, agen, dan sinkronisasi data inventaris.
- Audit stock opname, validasi kesesuaian fisik vs sistem (GR/DC), pengelolaan retur, dan moving stock.
- Evaluasi indikator kinerja rantai pasok seperti Fulfillment Delivery dan On Time Arrival (OTA).
- Pengolahan data analitis berbasis Excel (Pivot Table, VLOOKUP), SQL/Python, serta penyusunan SOP & manual sistem.

Atas komitmen dan kinerja operasional tersebut, saya dianugerahi penghargaan 'Best SCM Indonesia 2025 - Plant Pasuruan'. Dengan latar belakang ini, saya yakin mampu memberikan kontribusi optimal dalam mendukung kelancaran operasional gudang, pengelolaan buffer stock, dan akurasi inventory di {perusahaan}.

Sebagai bahan pertimbangan, bersama email ini saya lampirkan berkas terpadu Surat Lamaran Kerja dan Curriculum Vitae (CV) dalam format PDF.

Saya bersedia untuk mengikuti tahapan interview secara onsite/offline maupun sistem kerja shift. Besar harapan saya untuk diberikan kesempatan wawancara agar dapat menjelaskan kualifikasi saya secara lebih mendalam.

Atas perhatian dan kesempatan yang Bapak/Ibu berikan, saya ucapkan terima kasih.

Hormat saya,
{nama_pelamar}
📞 {telepon} | ✉️ {email_pelamar}
📍 {alamat_pelamar}"""

      # A. Generate PDF Surat Lamaran
      html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <style>
                    @page {{ size: A4; margin: 15mm 20mm 15mm 20mm; }}
                    body {{ font-family: 'Times New Roman', serif; font-size: 11pt; line-height: 1.35; color: #000; }}
                    p {{ margin: 6px 0; }}
                    .header {{ text-align: right; margin-bottom: 12px; }}
                    .recipient {{ margin-bottom: 15px; }}
                    .content {{ text-align: justify; }}
                    .bio-table {{ width: 100%; margin-left: 15px; margin-top: 4px; margin-bottom: 8px; }}
                    .bio-table td {{ padding: 1px 0; vertical-align: top; }}
                    .bio-table td.label {{ width: 26%; }}
                    .bio-table td.colon {{ width: 3%; }}
                    ol {{ margin-top: 2px; margin-bottom: 6px; padding-left: 22px; }}
                    ol li {{ padding: 1px 0; }}
                    .signature-table {{ width: 100%; margin-top: 15px; }}
                </style>
            </head>
            <body>
                <div class="header">Surabaya, {tanggal}</div>
                <div class="recipient">Yth. Bapak/Ibu HRD<br><strong>{perusahaan}</strong><br>di {lokasi}</div>
                <div class="content">
                    <p>Dengan hormat,</p>
                    <p>Berdasarkan informasi lowongan pekerjaan yang saya peroleh, bahwa <strong>{perusahaan}</strong> sedang membuka lowongan pekerjaan. Melalui surat ini saya bermaksud menyampaikan ketertarikan saya untuk melamar pekerjaan pada posisi <strong>{posisi}</strong>.</p>
                    <p>Berikut adalah biodata singkat saya:</p>
                    <table class="bio-table">
                        <tr><td class="label">Nama</td><td class="colon">:</td><td>{nama_pelamar}</td></tr>
                        <tr><td class="label">Pendidikan Terakhir</td><td class="colon">:</td><td>{pendidikan}</td></tr>
                        <tr><td class="label">Alamat Domisili</td><td class="colon">:</td><td>{alamat_pelamar}</td></tr>
                        <tr><td class="label">No. Telepon / WA</td><td class="colon">:</td><td>{telepon}</td></tr>
                        <tr><td class="label">Email</td><td class="colon">:</td><td>{email_pelamar}</td></tr>
                    </table>
                    <p>Saat ini saya dalam kondisi kesehatan yang sangat baik dan siap untuk bekerja keras serta berkontribusi positif bagi perusahaan. Saya memiliki latar belakang pendidikan S2 Teknologi Informasi serta pengalaman kerja lebih dari 8 tahun di bidang SCM, Data Control, dan IT Support yang saya yakini dapat mendukung kinerja operasional pada posisi tersebut.</p>
                    <p>Sebagai bahan pertimbangan Bapak/Ibu, bersama surat ini turut saya lampirkan:</p>
                    <ol>
                        <li>Curriculum Vitae (CV)</li>
                        <li>Fotokopi Ijazah Terakhir dan Transkrip Nilai</li>
                        <li>Portofolio / Dokumen Pendukung</li>
                        <li>Pasfoto Terbaru</li>
                    </ol>
                    <p>Besar harapan saya agar Bapak/Ibu bersedia meluangkan waktu untuk memberikan kesempatan wawancara. Atas perhatian dan kesempatan yang diberikan, saya ucapkan terima kasih.</p>
                </div>
                <table class="signature-table">
                    <tr>
                        <td style="width: 60%;"></td>
                        <td style="width: 40%;">
                            <p style="margin:0;">Hormat saya,</p>
                            {dapatkan_html_ttd()}
                            <p style="text-decoration: underline; font-weight: bold; margin-top: 0;">{nama_pelamar}</p>
                        </td>
                    </tr>
                </table>
            </body>
            </html>
            """
      surat_buffer = BytesIO()
      pisa.CreatePDF(html_content, dest=surat_buffer)
      surat_buffer.seek(0)

      # B. Merge Surat + CV
      if PdfMerger is not None:
        merger = PdfMerger()
        merger.append(surat_buffer)

        if uploaded_cv is not None:
          merger.append(BytesIO(uploaded_cv.read()))
        elif os.path.exists("cv.pdf"):
          merger.append("cv.pdf")
        elif os.path.exists("CV Adithya Marhaendra Kusuma ats.pdf"):
          merger.append("CV Adithya Marhaendra Kusuma ats.pdf")

        final_buffer = BytesIO()
        merger.write(final_buffer)
        merger.close()
        final_pdf_data = final_buffer.getvalue()
      else:
        final_pdf_data = surat_buffer.getvalue()

      nama_pdf = f"Surat_Lamaran_dan_CV_{perusahaan.replace(' ', '_')}.pdf"

      # C. Rekap Ke Excel
      if os.path.exists(file_excel):
        wb = openpyxl.load_workbook(file_excel)
        ws = (
            wb["Daftar Lamaran"]
            if "Daftar Lamaran" in wb.sheetnames
            else wb.active
        )
      else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar Lamaran"
        ws.append([
            "Tanggal",
            "Nama Perusahaan",
            "Lokasi",
            "Posisi",
            "Status",
            "Nama File PDF",
        ])

      ws.append([
          tanggal,
          perusahaan,
          lokasi,
          posisi,
          "Draft Prepared",
          nama_pdf,
      ])
      wb.save(file_excel)

      # D. Simpan ke Session State untuk Preview
      st.session_state.pdf_ready = True
      st.session_state.final_pdf_data = final_pdf_data
      st.session_state.nama_pdf = nama_pdf
      st.session_state.email_tujuan = email_tujuan
      st.session_state.subjek_email = subjek_draft
      st.session_state.body_email = body_draft
      st.session_state.app_password = gmail_app_password

  # ================= AREA PREVIEW & PENGIRIMAN MANUAL =================
  if st.session_state.pdf_ready:
    st.divider()
    st.success(
        "✅ Dokumen PDF Berhasil Dibuat! Silakan periksa atau ubah draft email"
        " di bawah ini sebelum dikirim."
    )

    col_dl, col_blank = st.columns([1, 1])
    with col_dl:
      st.download_button(
          label="📥 Download & Cek File PDF",
          data=st.session_state.final_pdf_data,
          file_name=st.session_state.nama_pdf,
          mime="application/pdf",
          use_container_width=True,
      )

    st.subheader("📧 Form Preview & Edit Email")
    final_to = st.text_input(
        "Email Tujuan (HRD):", value=st.session_state.email_tujuan
    )
    final_subject = st.text_input(
        "Subjek Email:", value=st.session_state.subjek_email
    )
    final_body = st.text_area(
        "Isi Body Email (Bisa Diedit):",
        value=st.session_state.body_email,
        height=300,
    )

    if st.button("📤 Kirim Email Sekarang", type="primary"):
      if not st.session_state.app_password:
        st.error("Gmail App Password belum terisi!")
      else:
        try:
          msg = MIMEMultipart()
          msg["From"] = email_pelamar
          msg["To"] = final_to
          msg["Subject"] = final_subject
          msg.attach(MIMEText(final_body, "plain"))

          part = MIMEBase("application", "octet-stream")
          part.set_payload(st.session_state.final_pdf_data)
          encoders.encode_base64(part)
          part.add_header(
              "Content-Disposition",
              f"attachment; filename={st.session_state.nama_pdf}",
          )
          msg.attach(part)

          server = smtplib.SMTP("smtp.gmail.com", 587)
          server.starttls()
          server.login(
              email_pelamar, st.session_state.app_password.replace(" ", "")
          )
          server.sendmail(email_pelamar, final_to, msg.as_string())
          server.quit()

          st.balloons()
          st.success(
              f"🚀 Email & Berkas Lampiran PDF Berhasil Terkirim ke {final_to}!"
          )
        except Exception as e:
          st.error(f"Gagal mengirim email: {e}")

# ================= TAB 2: REKAP DATA =================
with tab2:
  st.subheader("📊 Histori Lamaran Kerja")
  if os.path.exists(file_excel):
    df = pd.read_excel(file_excel)
    st.metric(label="Total Lamaran Dibuat", value=len(df))
    st.dataframe(df, use_container_width=True)
  else:
    st.info("Belum ada data lamaran yang dibuat.")
